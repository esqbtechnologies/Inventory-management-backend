from genericpath import exists
from django.http import HttpResponse, JsonResponse
from requests import Response
import ivnt_mngmnt.settings as set
from rest_framework.views import APIView
from rest_framework.generics import ListAPIView
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework_jwt.authentication import JSONWebTokenAuthentication
import json
import jwt
from rest_framework import status
import pandas as pd
import os
from openpyxl import load_workbook
import xlsxwriter

class generateFar(APIView):
    permission_classes = (AllowAny,)
    authentication_classes = (JSONWebTokenAuthentication,)

    def post(self,request):
        excel = request.data['File']
        wb1 = load_workbook(excel)
        sheet_names1 = wb1.get_sheet_names()
        name1 = sheet_names1[0]
        sheet_ranges1 = wb1[name1]
        df= pd.DataFrame(sheet_ranges1.values)
        path = os.path.join(set.BASE_DIR, 'FAR Template-1.xlsx')
        wb = load_workbook(filename=path)
        sheet_names = wb.get_sheet_names()
        name = sheet_names[0]
        sheet_ranges = wb[name]
        df_excel = pd.DataFrame(sheet_ranges.values)
        for i in range(0,30):
            df_excel[i] = df[i]
        sheet_ranges2 = wb[sheet_names[1]]
        df_excel2 = pd.DataFrame(sheet_ranges2.values)    
        writer = pd.ExcelWriter('Output.xlsx',engine='xlsxwriter',datetime_format='YYYY-MM-DD HH:MM:SS')  
        df_excel.to_excel(writer,sheet_name= 'Input',header= False,index=False)
        df_excel2.to_excel(writer,sheet_name='FAR Report',header=False,index=False)
        # writer.save()
        response = HttpResponse(content=writer, content_type='application/ms-excel')
        return response
